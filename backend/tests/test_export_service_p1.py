from __future__ import annotations

from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

from app.services import export_service


def test_export_teacher_report_xlsx_has_required_sheets(monkeypatch):
    payload = {
        "students": [
            {
                "student_name": "Nguyễn Văn A",
                "student_id": 1,
                "placement_score": 45,
                "final_score": 72,
                "level": "Khá",
                "pass_fail": "Pass",
                "study_hours": 3.5,
                "topics_completed": 4,
                "exercises_done": 10,
            }
        ],
        "attempts": [SimpleNamespace(id=10, user_id=1, quiz_set_id=5, score_percent=72, duration_sec=900, created_at="2026-01-01")],
        "weak_topics": [{"topic_name": "Hàm số", "avg_score": 48.5, "student_count_struggling": 1}],
    }
    monkeypatch.setattr(export_service, "_collect_classroom_rows", lambda db, classroom_id: payload)

    blob = export_service.export_teacher_report_xlsx(db=None, classroom_id=1)
    wb = load_workbook(BytesIO(blob))
    assert wb.sheetnames == ["Điểm số", "Tiến độ", "Chủ đề yếu", "Raw Attempts"]
