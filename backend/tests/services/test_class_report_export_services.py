from io import BytesIO

import openpyxl

from app.services.excel_report_service import generate_class_report_excel
from app.services.pdf_report_service import generate_class_report_pdf


def _sample_students():
    return [
        {
            "name": "Nguyễn Văn A",
            "email": "a@example.com",
            "placement_score": 60,
            "final_score": 80,
            "level": "kha",
            "study_hours": 15,
            "ai_comment": "Tiến bộ tốt",
        },
        {
            "name": "Trần Thị B",
            "email": "b@example.com",
            "placement_score": 55,
            "final_score": 70,
            "level": "trung_binh",
            "study_hours": 12,
            "ai_comment": "Cần luyện thêm",
        },
    ]


def test_generate_class_report_pdf_returns_pdf_bytes():
    content = generate_class_report_pdf({"name": "Lớp 12A1"}, _sample_students())
    assert isinstance(content, bytes)
    assert content.startswith(b"%PDF")


def test_generate_class_report_excel_returns_workbook_bytes():
    content = generate_class_report_excel({"name": "Lớp 12A1"}, _sample_students())
    assert isinstance(content, bytes)

    wb = openpyxl.load_workbook(BytesIO(content))
    assert "Danh sách học sinh" in wb.sheetnames
    assert "Thống kê điểm" in wb.sheetnames
    assert wb["Danh sách học sinh"]["A1"].value == "STT"
