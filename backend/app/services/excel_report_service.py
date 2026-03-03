from __future__ import annotations

from io import BytesIO
from typing import Any

import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill


_LEVEL_MAP = {
    "gioi": "Giỏi",
    "kha": "Khá",
    "trung_binh": "Trung bình",
    "yeu": "Yếu",
}


def generate_class_report_excel(classroom_data: dict[str, Any], students_data: list[dict[str, Any]]) -> bytes:
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Danh sách học sinh"

    headers = [
        "STT",
        "Họ và tên",
        "Email",
        "Điểm đầu vào",
        "Điểm cuối kỳ",
        "Xếp loại",
        "Thời gian học (giờ)",
        "Nhận xét AI",
    ]
    header_fill = PatternFill("solid", fgColor="4A90D9")

    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_num, student in enumerate(students_data, 2):
        ws1.cell(row=row_num, column=1, value=row_num - 1)
        ws1.cell(row=row_num, column=2, value=student.get("name") or "N/A")
        ws1.cell(row=row_num, column=3, value=student.get("email") or "")
        ws1.cell(row=row_num, column=4, value=float(student.get("placement_score") or 0.0))
        ws1.cell(row=row_num, column=5, value=float(student.get("final_score") or 0.0))
        ws1.cell(row=row_num, column=6, value=_LEVEL_MAP.get(str(student.get("level") or "").lower(), student.get("level") or "N/A"))
        ws1.cell(row=row_num, column=7, value=float(student.get("study_hours") or 0.0))
        ws1.cell(row=row_num, column=8, value=student.get("ai_comment") or "")

    column_widths = {1: 8, 2: 30, 3: 28, 4: 16, 5: 16, 6: 14, 7: 18, 8: 45}
    for col, width in column_widths.items():
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    ws2 = wb.create_sheet("Thống kê điểm")
    ws2.append(["Học sinh", "Điểm đầu vào", "Điểm cuối kỳ"])
    for student in students_data:
        ws2.append([
            student.get("name") or "N/A",
            float(student.get("placement_score") or 0.0),
            float(student.get("final_score") or 0.0),
        ])

    chart = BarChart()
    chart.title = f"Phân bố điểm lớp: {classroom_data.get('name') or 'N/A'}"
    chart.y_axis.title = "Điểm"
    chart.x_axis.title = "Học sinh"

    data_ref = Reference(ws2, min_col=2, max_col=3, min_row=1, max_row=max(2, len(students_data) + 1))
    categories = Reference(ws2, min_col=1, min_row=2, max_row=max(2, len(students_data) + 1))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 7
    chart.width = 14
    ws2.add_chart(chart, "E2")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
