from __future__ import annotations

import os
import tempfile
from pathlib import Path
from typing import Any

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def _font_name() -> str:
    candidates = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSansCondensed.ttf",
    ]
    for path in candidates:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont("DejaVu", path))
                return "DejaVu"
            except Exception:
                continue
    return "Helvetica"


def _build_score_chart(students: list[dict[str, Any]]) -> str:
    import matplotlib.pyplot as plt
    fig, ax = plt.subplots(figsize=(6, 3))
    names = [str(s.get("name") or f"HS{i+1}") for i, s in enumerate(students[:10])]
    pre = [float(s.get("entry_score") or 0) for s in students[:10]]
    post = [float(s.get("final_score") or 0) for s in students[:10]]
    x = range(len(names))
    ax.plot(x, pre, marker="o", label="Pre")
    ax.plot(x, post, marker="o", label="Post")
    ax.set_xticks(list(x), names, rotation=30, ha="right")
    ax.set_ylabel("Score")
    ax.legend()
    ax.set_title("Tiến bộ Pre/Post")
    out = Path(tempfile.mkstemp(prefix="lms_chart_", suffix=".png")[1])
    fig.tight_layout()
    fig.savefig(out, dpi=160)
    plt.close(fig)
    return str(out)


def export_report_pdf(report: dict[str, Any], *, name: str = "report") -> str:
    out = Path(tempfile.mkstemp(prefix=f"{name}_", suffix=".pdf")[1])
    font_name = _font_name()
    styles = getSampleStyleSheet()
    normal = ParagraphStyle("vn", parent=styles["Normal"], fontName=font_name, fontSize=10)
    heading = ParagraphStyle("hd", parent=styles["Heading1"], fontName=font_name)

    doc = SimpleDocTemplate(str(out), pagesize=A4)
    story: list[Any] = [Paragraph("Báo cáo học tập", heading), Spacer(1, 8)]

    students = report.get("students") or report.get("student_reports") or []
    try:
        chart = _build_score_chart(students)
        story.append(Image(chart, width=440, height=220))
        story.append(Spacer(1, 10))
    except Exception:
        story.append(Paragraph("(Không thể render chart trong môi trường hiện tại)", normal))
        story.append(Spacer(1, 10))

    table_data = [["Học viên", "Pre", "Post", "Mức", "Giờ học"]]
    for s in students:
        table_data.append([
            str(s.get("name") or s.get("student_name") or "N/A"),
            f"{float(s.get('entry_score', s.get('pre_score', 0)) or 0):.1f}",
            f"{float(s.get('final_score', s.get('post_score', 0)) or 0):.1f}",
            str(s.get("level") or s.get("current_level") or ""),
            f"{float(s.get('study_time_minutes', 0) or 0)/60.0:.2f}",
        ])

    t = Table(table_data, repeatRows=1)
    t.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
    ]))
    story.append(t)
    story.append(Spacer(1, 10))
    story.append(Paragraph("Biểu đồ và bảng trên dùng font Unicode tiếng Việt để tránh lỗi vỡ chữ.", normal))

    doc.build(story)
    return str(out)


def export_report_xlsx(report: dict[str, Any], *, name: str = "report") -> str:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Gradebook"
    ws.append(["student_id", "name", "pre", "mid", "post", "avg", "level"]) 

    students = report.get("students") or report.get("student_reports") or []
    for s in students:
        pre = float(s.get("entry_score", s.get("pre_score", 0)) or 0)
        mid = float(s.get("mid_score", 0) or 0)
        post = float(s.get("final_score", s.get("post_score", 0)) or 0)
        avg = (pre + mid + post) / 3.0
        ws.append([s.get("student_id"), s.get("name") or s.get("student_name"), pre, mid, post, avg, s.get("level") or s.get("current_level")])

    ws2 = wb.create_sheet("TopicBreakdown")
    ws2.append(["student_id", "topic", "score"])
    for s in students:
        sid = s.get("student_id")
        for t, v in (s.get("topic_scores") or {}).items():
            ws2.append([sid, str(t), float(v or 0)])

    ws3 = wb.create_sheet("StudyTime")
    ws3.append(["student_id", "minutes", "hours"])
    for s in students:
        m = float(s.get("study_time_minutes", 0) or 0)
        ws3.append([s.get("student_id"), m, m / 60.0])

    ws4 = wb.create_sheet("Summary")
    ws4.append(["metric", "value"])
    summary = report.get("summary") or report.get("class_summary") or {}
    for k, v in summary.items():
        ws4.append([k, str(v)])

    out = Path(tempfile.mkstemp(prefix=f"{name}_", suffix=".xlsx")[1])
    wb.save(str(out))
    return str(out)
