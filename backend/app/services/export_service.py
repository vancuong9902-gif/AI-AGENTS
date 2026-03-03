from __future__ import annotations

from io import BytesIO
from pathlib import Path
from statistics import mean
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.orm import Session

from app.models.attempt import Attempt
from app.models.classroom import Classroom, ClassroomMember
from app.models.document_topic import DocumentTopic
from app.models.quiz_set import QuizSet
from app.models.user import User
from app.services.lms_service import classify_student_level


def _font_path() -> str | None:
    candidates = [
        Path("static/fonts/NotoSans-Regular.ttf"),
        Path("backend/static/fonts/NotoSans-Regular.ttf"),
    ]
    for c in candidates:
        if c.exists():
            return str(c)
    return None


def _register_font() -> str:
    fp = _font_path()
    if fp:
        try:
            pdfmetrics.registerFont(TTFont("NotoSans", fp))
            return "NotoSans"
        except Exception:
            pass
    return "Helvetica"


def _collect_classroom_rows(db: Session, classroom_id: int) -> dict[str, Any]:
    member_rows = db.query(ClassroomMember.user_id).filter(ClassroomMember.classroom_id == int(classroom_id)).all()
    student_ids = [int(r[0]) for r in member_rows]
    users = db.query(User).filter(User.id.in_(student_ids)).all() if student_ids else []
    by_user = {int(u.id): u for u in users}

    attempts = (
        db.query(Attempt)
        .join(QuizSet, QuizSet.id == Attempt.quiz_set_id)
        .filter(Attempt.user_id.in_(student_ids))
        .order_by(Attempt.created_at.desc())
        .all()
        if student_ids
        else []
    )

    latest_pre: dict[int, Attempt] = {}
    latest_final: dict[int, Attempt] = {}
    for a in attempts:
        q = db.query(QuizSet).filter(QuizSet.id == int(a.quiz_set_id)).first()
        kind = str(getattr(q, "kind", "") or "").lower()
        uid = int(a.user_id)
        if kind in {"diagnostic_pre", "entry_test"} and uid not in latest_pre:
            latest_pre[uid] = a
        if kind in {"diagnostic_post", "final_exam", "final"} and uid not in latest_final:
            latest_final[uid] = a

    students = []
    weak_topic_counter: dict[str, list[float]] = {}
    for uid in student_ids:
        pre = float(getattr(latest_pre.get(uid), "score_percent", 0) or 0)
        fin = float(getattr(latest_final.get(uid), "score_percent", 0) or 0)
        level = classify_student_level(int(round(fin if fin > 0 else pre)))
        row = {
            "student_id": uid,
            "student_name": str(getattr(by_user.get(uid), "full_name", "") or f"User #{uid}"),
            "placement_score": round(pre, 1),
            "final_score": round(fin, 1),
            "level": str(level.get("label") or "Trung Bình"),
            "pass_fail": "Pass" if fin >= 50 else "Fail",
            "study_hours": round((float(getattr(latest_final.get(uid), "duration_sec", 0) or 0) / 3600.0), 2),
            "topics_completed": 0,
            "exercises_done": 0,
        }
        students.append(row)

        br = getattr(latest_final.get(uid), "breakdown_json", []) or []
        for item in br:
            topic = str(item.get("topic") or "General").strip() or "General"
            max_points = float(item.get("max_points") or 0)
            score_points = float(item.get("score_points") or 0)
            pct = (score_points / max_points * 100.0) if max_points > 0 else 0.0
            weak_topic_counter.setdefault(topic, []).append(pct)

    weak_topics = sorted(
        [
            {
                "topic_name": topic,
                "avg_score": round(mean(scores), 1),
                "student_count_struggling": sum(1 for x in scores if x < 60),
            }
            for topic, scores in weak_topic_counter.items()
        ],
        key=lambda x: x["avg_score"],
    )[:5]

    return {
        "students": students,
        "attempts": attempts,
        "weak_topics": weak_topics,
    }


def export_teacher_report_pdf(db: Session, classroom_id: int) -> bytes:
    classroom = db.query(Classroom).filter(Classroom.id == int(classroom_id)).first()
    payload = _collect_classroom_rows(db, int(classroom_id))
    students = payload["students"]
    weak_topics = payload["weak_topics"]

    font_name = _register_font()
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("vn-title", parent=styles["Title"], fontName=font_name)
    body_style = ParagraphStyle("vn-body", parent=styles["Normal"], fontName=font_name)

    elems = [
        Paragraph("BÁO CÁO LỚP HỌC", title_style),
        Spacer(1, 10),
        Paragraph(f"Lớp: {getattr(classroom, 'name', f'#{classroom_id}')}", body_style),
        Spacer(1, 6),
        Paragraph("Thống kê điểm học viên", body_style),
        Spacer(1, 10),
    ]

    score_rows = [["Học viên", "Placement", "Final", "Cấp độ", "Giờ học"]]
    for s in students:
        score_rows.append([s["student_name"], s["placement_score"], s["final_score"], s["level"], s["study_hours"]])

    t = Table(score_rows, repeatRows=1)
    t.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E5E7EB")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elems.append(t)
    elems.append(Spacer(1, 12))

    weak_rows = [["Top chủ đề yếu", "Điểm TB", "Số HS yếu"]]
    for w in weak_topics:
        weak_rows.append([w["topic_name"], w["avg_score"], w["student_count_struggling"]])
    if len(weak_rows) == 1:
        weak_rows.append(["N/A", 0, 0])

    wt = Table(weak_rows, repeatRows=1)
    wt.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FEF3C7")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elems.append(wt)

    doc.build(elems)
    return buf.getvalue()


def export_teacher_report_xlsx(db: Session, classroom_id: int) -> bytes:
    payload = _collect_classroom_rows(db, int(classroom_id))
    students = payload["students"]
    attempts = payload["attempts"]
    weak_topics = payload["weak_topics"]

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Điểm số"
    ws1.append(["student_name", "student_code", "placement_score", "final_score", "level", "pass/fail"])
    for s in students:
        ws1.append([s["student_name"], s["student_id"], s["placement_score"], s["final_score"], s["level"], s["pass_fail"]])

    ws2 = wb.create_sheet("Tiến độ")
    ws2.append(["student_name", "topics_completed", "exercises_done", "study_hours", "last_active"])
    for s in students:
        ws2.append([s["student_name"], s["topics_completed"], s["exercises_done"], s["study_hours"], ""])

    ws3 = wb.create_sheet("Chủ đề yếu")
    ws3.append(["topic_name", "avg_score", "student_count_struggling"])
    for w in weak_topics:
        ws3.append([w["topic_name"], w["avg_score"], w["student_count_struggling"]])

    ws4 = wb.create_sheet("Raw Attempts")
    ws4.append(["attempt_id", "student", "quiz_type", "score", "time_spent", "submitted_at"])
    for a in attempts:
        ws4.append([int(a.id), int(a.user_id), int(a.quiz_set_id), float(a.score_percent or 0), int(a.duration_sec or 0), str(a.created_at or "")])

    header_fill = PatternFill("solid", fgColor="E5E7EB")
    for ws in [ws1, ws2, ws3, ws4]:
        for c in ws[1]:
            c.font = Font(bold=True)
            c.fill = header_fill
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(35, max(12, max(len(str(cell.value or "")) for cell in col) + 2))

    red = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    yellow = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    green = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    ws1.conditional_formatting.add("D2:D999", CellIsRule(operator="lessThan", formula=["50"], fill=red))
    ws1.conditional_formatting.add("D2:D999", CellIsRule(operator="between", formula=["50", "70"], fill=yellow))
    ws1.conditional_formatting.add("D2:D999", CellIsRule(operator="greaterThan", formula=["70"], fill=green))

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()
