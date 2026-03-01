from __future__ import annotations

import math
import tempfile
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from statistics import mean, pstdev

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.models.attempt import Attempt
from app.models.classroom import Classroom, ClassroomMember
from app.models.classroom_assessment import ClassroomAssessment
from app.models.learner_profile import LearnerProfile
from app.models.quiz_set import QuizSet
from app.models.session import Session as UserSession
from app.models.user import User
from app.services.lms_service import classify_student_level


def _safe_float(value: object) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except Exception:
        return None


def _get_topic_mastery_map(profile: LearnerProfile | None) -> dict[str, float | None]:
    if not profile or not isinstance(profile.mastery_json, dict):
        return {}

    mastery_payload = profile.mastery_json
    raw_topics = mastery_payload.get("topic_mastery")
    if not isinstance(raw_topics, dict):
        return {}

    topic_map: dict[str, float | None] = {}
    for topic, value in raw_topics.items():
        topic_name = str(topic or "").strip()
        if not topic_name:
            continue
        numeric = _safe_float(value)
        if numeric is None:
            topic_map[topic_name] = None
            continue
        topic_map[topic_name] = max(0.0, min(100.0, numeric))

    return topic_map


def _sheet_autofit(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        col = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col].width = min(max(10, max_length + 2), 60)


def _apply_score_conditional_formatting(ws, *, start_row: int, end_row: int, columns: list[str]) -> None:
    if end_row < start_row:
        return

    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    orange = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for col in columns:
        rng = f"{col}{start_row}:{col}{end_row}"
        ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThanOrEqual", formula=["85"], fill=green))
        ws.conditional_formatting.add(rng, CellIsRule(operator="between", formula=["70", "84.9999"], fill=yellow))
        ws.conditional_formatting.add(rng, CellIsRule(operator="between", formula=["50", "69.9999"], fill=orange))
        ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["50"], fill=red))


def _latest_score(db: Session, *, user_id: int, kind: str) -> float | None:
    row = (
        db.query(Attempt)
        .join(QuizSet, QuizSet.id == Attempt.quiz_set_id)
        .filter(Attempt.user_id == int(user_id), QuizSet.kind == str(kind))
        .order_by(Attempt.created_at.desc(), Attempt.id.desc())
        .first()
    )
    return _safe_float(getattr(row, "score_percent", None)) if row else None


def export_classroom_gradebook_xlsx(db: Session, classroom_id: int) -> Path:
    classroom = db.query(Classroom).filter(Classroom.id == int(classroom_id)).first()
    if not classroom:
        raise ValueError(f"Classroom {classroom_id} not found")

    students = (
        db.query(User)
        .join(ClassroomMember, ClassroomMember.user_id == User.id)
        .filter(ClassroomMember.classroom_id == int(classroom_id))
        .order_by(User.full_name.asc().nulls_last(), User.id.asc())
        .all()
    )

    assigned_midterm_ids = {
        int(row[0])
        for row in (
            db.query(ClassroomAssessment.assessment_id)
            .join(QuizSet, QuizSet.id == ClassroomAssessment.assessment_id)
            .filter(
                ClassroomAssessment.classroom_id == int(classroom_id),
                QuizSet.kind == "midterm",
            )
            .all()
        )
    }

    student_ids = [int(s.id) for s in students]
    session_hours_map: dict[int, float] = {}
    if student_ids:
        session_rows = (
            db.query(UserSession.user_id, UserSession.started_at, UserSession.ended_at)
            .filter(UserSession.user_id.in_(student_ids))
            .all()
        )
        has_completed_session = any(r.ended_at is not None and r.started_at is not None for r in session_rows)
        if has_completed_session:
            for r in session_rows:
                if not r.ended_at or not r.started_at:
                    continue
                started_at = r.started_at
                ended_at = r.ended_at
                if started_at.tzinfo is None:
                    started_at = started_at.replace(tzinfo=timezone.utc)
                if ended_at.tzinfo is None:
                    ended_at = ended_at.replace(tzinfo=timezone.utc)
                duration_hours = max(0.0, (ended_at - started_at).total_seconds() / 3600.0)
                session_hours_map[int(r.user_id)] = session_hours_map.get(int(r.user_id), 0.0) + duration_hours

    attempt_hours_map: dict[int, float] = {}
    if student_ids:
        for uid, seconds_sum in (
            db.query(Attempt.user_id, func.coalesce(func.sum(Attempt.duration_sec), 0))
            .filter(Attempt.user_id.in_(student_ids))
            .group_by(Attempt.user_id)
            .all()
        ):
            attempt_hours_map[int(uid)] = float(seconds_sum or 0.0) / 3600.0

    profile_map = {
        int(p.user_id): p
        for p in db.query(LearnerProfile).filter(LearnerProfile.user_id.in_(student_ids)).all()
    }

    wb = Workbook()
    ws_gradebook = wb.active
    ws_gradebook.title = "Tong hop diem"
    ws_gradebook.freeze_panes = "A2"
    ws_gradebook.append([
        "STT",
        "Student ID",
        "Ho ten",
        "Email",
        "Diem dau vao (pre)",
        "Diem cuoi ky (post)",
        "TB bai tap (midterm avg)",
        "Tong gio hoc",
        "Xep loai",
    ])

    topic_union: set[str] = set()
    student_topic_map: dict[int, dict[str, float | None]] = {}

    metric_pre: list[float] = []
    metric_post: list[float] = []
    metric_midterm: list[float] = []
    level_counter: Counter[str] = Counter()

    for index, student in enumerate(students, start=1):
        uid = int(student.id)
        pre_score = _latest_score(db, user_id=uid, kind="diagnostic_pre")
        post_score = _latest_score(db, user_id=uid, kind="diagnostic_post")

        midterm_query = db.query(Attempt).join(QuizSet, QuizSet.id == Attempt.quiz_set_id).filter(
            Attempt.user_id == uid,
            QuizSet.kind == "midterm",
        )
        if assigned_midterm_ids:
            midterm_query = midterm_query.filter(Attempt.quiz_set_id.in_(assigned_midterm_ids))
        midterm_scores = [float(a.score_percent or 0.0) for a in midterm_query.all()]
        midterm_avg = mean(midterm_scores) if midterm_scores else None

        study_hours = session_hours_map.get(uid)
        if study_hours is None:
            study_hours = attempt_hours_map.get(uid, 0.0)

        level_source = post_score if post_score is not None else (midterm_avg if midterm_avg is not None else (pre_score or 0.0))
        level_obj = classify_student_level(int(round(level_source)))
        level_label = str(level_obj.get("label") or level_obj.get("level_key") or "")
        level_counter[level_label] += 1

        profile = profile_map.get(uid)
        topic_mastery = _get_topic_mastery_map(profile)
        student_topic_map[uid] = topic_mastery
        topic_union.update(topic_mastery.keys())

        ws_gradebook.append([
            index,
            uid,
            student.full_name or f"Hoc sinh {uid}",
            student.email,
            round(pre_score, 2) if pre_score is not None else None,
            round(post_score, 2) if post_score is not None else None,
            round(midterm_avg, 2) if midterm_avg is not None else None,
            round(float(study_hours or 0.0), 2),
            level_label,
        ])

        if pre_score is not None:
            metric_pre.append(float(pre_score))
        if post_score is not None:
            metric_post.append(float(post_score))
        if midterm_avg is not None:
            metric_midterm.append(float(midterm_avg))

    _sheet_autofit(ws_gradebook)
    _apply_score_conditional_formatting(ws_gradebook, start_row=2, end_row=ws_gradebook.max_row, columns=["E", "F", "G"])

    ws_topic = wb.create_sheet("Topic mastery")
    ws_topic.freeze_panes = "A2"
    ordered_topics = sorted(topic_union)
    ws_topic.append(["Student ID", "Ho ten", *ordered_topics])
    for student in students:
        uid = int(student.id)
        mastery_map = student_topic_map.get(uid, {})
        row = [uid, student.full_name or f"Hoc sinh {uid}"]
        for topic in ordered_topics:
            mastery_value = mastery_map.get(topic)
            row.append(round(mastery_value, 2) if mastery_value is not None else None)
        ws_topic.append(row)
    _sheet_autofit(ws_topic)

    ws_stats = wb.create_sheet("Thong ke lop")
    ws_stats.freeze_panes = "A2"
    ws_stats.append(["Chi so", "Pre", "Post", "Midterm"])

    def _summary(values: list[float]) -> tuple[float | None, float | None, float | None, float | None]:
        if not values:
            return None, None, None, None
        avg_v = mean(values)
        min_v = min(values)
        max_v = max(values)
        std_v = pstdev(values) if len(values) > 1 else 0.0
        return avg_v, min_v, max_v, std_v

    stats_rows = [
        ("avg", *_summary(metric_pre)[0:1], *_summary(metric_post)[0:1], *_summary(metric_midterm)[0:1]),
        ("min", *_summary(metric_pre)[1:2], *_summary(metric_post)[1:2], *_summary(metric_midterm)[1:2]),
        ("max", *_summary(metric_pre)[2:3], *_summary(metric_post)[2:3], *_summary(metric_midterm)[2:3]),
        ("std", *_summary(metric_pre)[3:4], *_summary(metric_post)[3:4], *_summary(metric_midterm)[3:4]),
    ]
    for stat_name, pre_val, post_val, mid_val in stats_rows:
        ws_stats.append([
            stat_name,
            round(pre_val, 2) if pre_val is not None and not math.isnan(pre_val) else None,
            round(post_val, 2) if post_val is not None and not math.isnan(post_val) else None,
            round(mid_val, 2) if mid_val is not None and not math.isnan(mid_val) else None,
        ])

    ws_stats.append([])
    ws_stats.append(["Xep loai", "So hoc sinh"])
    for level_label, count in sorted(level_counter.items()):
        ws_stats.append([level_label, int(count)])

    _sheet_autofit(ws_stats)
    _apply_score_conditional_formatting(ws_stats, start_row=2, end_row=5, columns=["B", "C", "D"])

    out_path = Path(tempfile.mkstemp(prefix=f"diem_lop_{classroom_id}_", suffix=".xlsx")[1])
    wb.save(str(out_path))
    return out_path
